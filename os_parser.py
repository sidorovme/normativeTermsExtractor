import openpyxl
import json
import re
from typing import Dict, List

def extract_text_and_notes(cell) -> tuple:
    if not cell or cell.value is None:
        return None, []

    if isinstance(cell.value, openpyxl.cell.rich_text.CellRichText):
        term = ""
        links_str = ""

        for part in cell.value:
            if hasattr(part, 'font'):
                if part.font.vertAlign == 'superscript': # only for upper text
                    #print(f"Part: {part}, Font: {part.font}")
                    links_str = links_str + part.text
                else:
                    term = term + part.text
            else:
                term = term + part # concatenating parts without decoration

        # Возвращаем None вместо пустой строки для term
        if not term.strip():
            term = None

        # Создаем список ссылок, только если они есть
        links = [item.strip() for item in links_str.split(",") if item.strip()]
        return term, links
    else:
        # Возвращаем None вместо пустой строки
        return None if not cell.value else cell.value, []


def parse_excel_to_json(file_path: str) -> Dict:
    """
    Анализирует Excel-файл и преобразует его в структурированный JSON.

    Результат содержит два корневых объекта:
    1. normativeTerms - иерархическая структура нормативных сроков
    2. notes - примечания из раздела "Примечание"

    Функция пропускает шапку и разделяет основной контент и примечания.
    """
    # Загружаем рабочую книгу
    workbook = openpyxl.load_workbook(file_path, rich_text=True)
    sheet = workbook["Нормативные сроки"]

    # Определяем границы для обработки (пропускаем шапку и разделяем основной контент и примечания)
    header_rows = 8  # Количество строк в шапке (включая строку с названиями столбцов)

    # Находим строку с началом примечаний
    note_row = None
    for row_idx in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=1).value
        if cell_value and "Примечание" in str(cell_value):
            note_row = row_idx
            break

    # Определяем диапазоны строк для обработки
    start_row = header_rows + 1  # Начинаем после шапки
    end_row = note_row - 1 if note_row else sheet.max_row  # Заканчиваем перед примечаниями

    # Результирующая структура данных
    result = {
        "normativeTerms": [],
        "notes": []
    }

    # Парсим основную часть документа (нормативные сроки)
    parse_normative_terms(sheet, start_row, end_row, result["normativeTerms"])

    # Парсим раздел примечаний, если он найден
    if note_row:
        parse_notes(sheet, note_row, sheet.max_row, result["notes"])

    return result


def parse_normative_terms(sheet, start_row: int, end_row: int, root_list: List):
    """
    Парсит основную часть документа с нормативными сроками.
    """
    # Текущие узлы для каждого уровня иерархии
    current_nodes = {
        1: None,  # Текущий узел уровня 1
        2: None,  # Текущий узел уровня 2
        3: None,  # Текущий узел уровня 3
        "group": None  # Текущий заголовок группы
    }

    # Обходим строки листа в определенном диапазоне
    for row_idx in range(start_row, end_row + 1):
        # Получаем данные из ячеек
        description_cell = sheet.cell(row=row_idx, column=1)
        #description = str(description_cell.value) if description_cell.value else ""

        caption, caption_notes = extract_text_and_notes(description_cell)
        description_node = {
            "value": caption
        }
        if caption_notes != []:
            description_node["notes"] = caption_notes

        code_cell = sheet.cell(row=row_idx, column=2)
        code = code_cell.value
        term_cell = sheet.cell(row=row_idx, column=3)

        # Пропускаем пустые строки
        if not caption and not code and (not term_cell.value or term_cell.value == ''):
            continue

        # Получаем отступ первой ячейки
        indent = description_cell.alignment.indent if hasattr(description_cell.alignment, 'indent') else 0

        # Извлекаем срок эксплуатации и ссылки
        term, term_notes = extract_text_and_notes(term_cell)
        term_node = {}
        if term is not None:
            term_node["value"] = term
        if term_notes != []:
            term_node["notes"] = term_notes

        # Создаем узел для текущей строки с обязательными полями
        node = {
            "description": description_node,
            "code": code
        }

        # Добавляем term только если он не null
        if term_node != {}:
            node["term"] = term_node

        # Если у узла есть потомки, добавляем соответствующее поле
        if not term and not code and caption:
            node["children"] = []

        # Определяем уровень иерархии
        if code and str(code).isdigit() and len(str(code)) == 1:
            # Уровень 1: шифр из 1 цифры
            node["children"] = []
            current_nodes[1] = node
            current_nodes[2] = None
            current_nodes[3] = None
            current_nodes["group"] = None
            root_list.append(node)

        elif code and str(code).isdigit() and len(str(code)) == 3:
            # Уровень 2: шифр из 3 цифр
            node["children"] = []
            current_nodes[2] = node
            current_nodes[3] = None
            current_nodes["group"] = None

            # Добавляем узел в дерево
            if current_nodes[1]:
                current_nodes[1]["children"].append(node)
            else:
                root_list.append(node)

        elif code and str(code).isdigit() and len(str(code)) == 5 and indent == 0:
            # Уровень 3: шифр из 5 цифр без отступа
            current_nodes[3] = node
            current_nodes["group"] = None

            # Добавляем узел в дерево
            if current_nodes[2]:
                current_nodes[2]["children"].append(node)
            elif current_nodes[1]:
                current_nodes[1]["children"].append(node)
            else:
                root_list.append(node)

        elif not code and not term and caption:
            # Уровень 3: заголовок группы
            node["children"] = []
            current_nodes["group"] = node

            # Добавляем узел в дерево
            if current_nodes[2]:
                current_nodes[2]["children"].append(node)
            elif current_nodes[1]:
                current_nodes[1]["children"].append(node)
            else:
                root_list.append(node)

        elif indent == 2:
            # Уровень 4: элемент с отступом 2
            # Добавляем к последнему заголовку группы
            if current_nodes["group"]:
                current_nodes["group"]["children"].append(node)
            elif current_nodes[3]:
                # Если нет заголовка группы, добавляем к последнему узлу уровня 3
                if "children" not in current_nodes[3]:
                    current_nodes[3]["children"] = []
                current_nodes[3]["children"].append(node)
            elif current_nodes[2]:
                current_nodes[2]["children"].append(node)
            elif current_nodes[1]:
                current_nodes[1]["children"].append(node)
            else:
                root_list.append(node)

    # Очищаем узлы от пустых children
    for node in root_list:
        clean_empty_children(node)


def parse_notes(sheet, start_row: int, end_row: int, notes_list: List):
    """
    Парсит раздел примечаний.
    Ключи создаются для строк, начинающихся с числа.
    Если строка не начинается с числа, ее содержимое добавляется к предыдущему примечанию.
    """
    current_note = None

    # Пропускаем первую строку с заголовком "Примечание"
    for row_idx in range(start_row + 1, end_row + 1):
        cell = sheet.cell(row=row_idx, column=1)
        if not cell.value:
            continue

        # Очищаем текст от непечатных символов
        text = str(cell.value).strip()
        text = re.sub(r'[\xa0\s]+', ' ', text).strip()  # Заменяем неразрывные пробелы и последовательности пробелов на один пробел

        if not text:
            continue

        # Проверяем, начинается ли строка с числа
        match = re.match(r'^\d+', text)

        if match:
            # Нашли новое примечание, начинающееся с числа
            key = match.group(0)
            # Остаток текста после числа - это содержимое примечания
            note_text = text[len(key):].strip()

            # Создаем новую запись примечания
            current_note = {"key": key, "note": note_text}
            notes_list.append(current_note)
        elif current_note:
            # Если строка не начинается с числа и у нас уже есть примечание,
            # добавляем текст к существующему примечанию
            current_note["note"] += " " + text


def clean_empty_children(node: Dict):
    """Удаляет пустые списки children из узлов"""
    if "children" in node and not node["children"]:
        del node["children"]

    if "children" in node:
        for child in node["children"]:
            clean_empty_children(child)


# Очистка строк от непечатных символов и неразрывных пробелов
def clean_string(text):
    if not text or not isinstance(text, str):
        return text
    # Заменяем непечатные символы и последовательности пробелов на один пробел
    return re.sub(r'[\xa0\s]+', ' ', text).strip()

# Рекурсивно очищаем все строки в объекте данных
def clean_data(obj):
    if isinstance(obj, dict):
        # Создаем список ключей для удаления (чтобы избежать изменения словаря во время итерации)
        keys_to_remove = []

        for key, value in obj.items():
            if isinstance(value, str):
                obj[key] = clean_string(value)
            elif isinstance(value, (dict, list)):
                clean_data(value)
            # Помечаем для удаления пустые списки и None значения
            elif value is None or (isinstance(value, list) and len(value) == 0):
                keys_to_remove.append(key)

        # Удаляем помеченные ключи
        for key in keys_to_remove:
            del obj[key]

    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            if isinstance(item, str):
                obj[i] = clean_string(item)
            elif isinstance(item, (dict, list)):
                clean_data(item)

# Пример использования
if __name__ == "__main__":
    file_path = "os.xlsx"
    result = parse_excel_to_json(file_path)

    # Очищаем все строки в результирующем объекте от непечатных символов
    clean_data(result)

    # Сохранение результата в JSON-файл с обработкой непечатных символов
    with open("result.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print("Данные успешно преобразованы в JSON и сохранены в файл result.json")
